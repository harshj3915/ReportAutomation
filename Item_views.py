import json
import os
import pandas as pd
import dash
from dash import dcc, html, dash_table, callback
from dash.dependencies import Input, Output, State
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import numpy as np

import os
import pandas as pd
import glob
from openpyxl import load_workbook
import time

# Define paths for saved data
cache_folder = os.path.join(os.path.dirname(__file__), 'cache', 'individual_files')
combined_cache_file = os.path.join(os.path.dirname(__file__), 'cache', 'combined_data.pkl')

# Function to get cache file path for a specific file
def get_cache_path(file_path):
    """Get cache file path for a specific data file"""
    file_name = os.path.basename(file_path)
    file_name_without_ext = os.path.splitext(file_name)[0]
    return os.path.join(cache_folder, f"{file_name_without_ext}.pkl")

# Function to check if cached file is newer than source file
def is_cache_valid(source_file, cache_file):
    """Check if cache file exists and is newer than source file"""
    if not os.path.exists(cache_file):
        return False
    
    source_mtime = os.path.getmtime(source_file)
    cache_mtime = os.path.getmtime(cache_file)
    return cache_mtime > source_mtime

# Function to clean orphaned cache files
def clean_orphaned_cache(source_files):
    """Remove cache files that no longer have corresponding source files"""
    if not os.path.exists(cache_folder):
        return
    
    # Get all cache files
    cache_files = glob.glob(os.path.join(cache_folder, '*.pkl'))
    source_basenames = [os.path.splitext(os.path.basename(f))[0] for f in source_files]
    
    for cache_file in cache_files:
        cache_basename = os.path.splitext(os.path.basename(cache_file))[0]
        if cache_basename not in source_basenames:
            try:
                os.remove(cache_file)
                print(f"Removed orphaned cache file: {os.path.basename(cache_file)}")
            except Exception as e:
                print(f"Error removing cache file {cache_file}: {e}")

# Check if combined cache exists first
# if os.path.exists(combined_cache_file):
#     print(f"\nCombined cache file found at {combined_cache_file}")
#     print(f"Loading cached data instead of processing files...")
#     combined_df = pd.read_pickle(combined_cache_file)
#     print("Cached data loaded successfully!")
    
#     # Filter out specific categories
#     categories_to_exclude = ['JSP', 'Remove', 'FOC-R', 'EMI', '(blank)', 'PRO', 'WRT', 'Others']
#     if 'Category' in combined_df.columns:
#         print(f"Filtering out categories: {', '.join(categories_to_exclude)}")
#         before_count = len(combined_df)
#         combined_df = combined_df[~combined_df['Category'].isin(categories_to_exclude)]
#         after_count = len(combined_df)
#         print(f"Removed {before_count - after_count} rows with excluded categories")
    
#     # Process the date column to extract day, month, and year
#     if 'Date' in combined_df.columns and not all(col in combined_df.columns for col in ['Day', 'Month', 'Year']):
#         print("Processing date column to extract day, month, and year...")
#         # Convert to datetime if it's not already
#         if not pd.api.types.is_datetime64_dtype(combined_df['Date']):
#             combined_df['Date'] = pd.to_datetime(combined_df['Date'], errors='coerce')
        
#         # Extract day, month and year
#         combined_df['Day'] = combined_df['Date'].dt.day
#         combined_df['Month'] = combined_df['Date'].dt.month_name()
#         combined_df['Year'] = combined_df['Date'].dt.year
        
#         # Save the updated dataframe back to cache
#         combined_df.to_pickle(combined_cache_file)
#         print("Date columns added and cache updated")
    
#     print("\nCombined data shape:", combined_df.shape)
#     print("\nCombined data columns:", combined_df.columns.tolist())
    
#     # Display the first few rows of the combined data
#     print("\nFirst few rows of combined data:")
#     print(combined_df.head())
    

print("No combined cache found. Processing files with individual caching...")

try:
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')
    print(f"Loading config from: {config_path}")
    
    with open(config_path, 'r') as f:
        config_data = json.load(f)
    folder_path = config_data['paths']['MainDashboardData']
    
    # Convert relative path to absolute path if needed
    if not os.path.isabs(folder_path):
        # If it's a relative path, make it relative to the current script directory
        script_dir = os.path.dirname(os.path.abspath(__file__))
        folder_path = os.path.join(script_dir, folder_path)
        folder_path = os.path.normpath(folder_path)  # Normalize the path
        
except Exception as e:
    print(f"❌ Error loading config.json: {e}")
    print("Using default folder path relative to script location...")
    script_dir = os.path.dirname(os.path.abspath(__file__))
    folder_path = os.path.join(script_dir, "q", "MainDashboardData")
    folder_path = os.path.normpath(folder_path)  # Normalize the path

print(f"Folder path: {folder_path}")
print(f"Folder exists: {os.path.exists(folder_path)}")

# Additional debugging - list contents if folder exists
if os.path.exists(folder_path):
    try:
        contents = os.listdir(folder_path)
        print(f"Contents of folder: {contents}")
        # Filter for Excel and CSV files specifically
        excel_csv_files = [f for f in contents if f.endswith(('.xlsx', '.csv'))]
        print(f"Excel/CSV files in folder: {excel_csv_files}")
    except Exception as e:
        print(f"Error listing folder contents: {e}")
else:
    # Try to find where the MainDashboardData folder might be
    script_dir = os.path.dirname(os.path.abspath(__file__))
    print(f"Script directory: {script_dir}")
    
    # Check if MainDashboardData exists in script directory
    possible_paths = [
        os.path.join(script_dir, "MainDashboardData"),
        os.path.join(script_dir, "q", "MainDashboardData"),
        os.path.join(script_dir, "..", "q", "MainDashboardData")
    ]
    
    for path in possible_paths:
        norm_path = os.path.normpath(path)
        print(f"Checking possible path: {norm_path} - Exists: {os.path.exists(norm_path)}")
        if os.path.exists(norm_path):
            print(f"Found data folder at: {norm_path}")
            folder_path = norm_path
            break

# Get a list of all Excel and CSV files in the folder
excel_files = glob.glob(os.path.join(folder_path, '*.xlsx'))
csv_files = glob.glob(os.path.join(folder_path, '*.csv'))
all_files = excel_files + csv_files
print(f"Found {len(excel_files)} Excel files and {len(csv_files)} CSV files (Total: {len(all_files)} files)")

# Clean orphaned cache files
clean_orphaned_cache(all_files)

# Debug: Print the actual files found
if excel_files:
    print("Excel files found:")
    for file in excel_files:
        print(f"  - {os.path.basename(file)}")
if csv_files:
    print("CSV files found:")
    for file in csv_files:
        print(f"  - {os.path.basename(file)}")

if not all_files:
    print(f"⚠️  No files found in folder: {folder_path}")
    print("Please check if the folder path is correct and contains .xlsx or .csv files")

# Function to get the sheet with the longest name from an Excel file
def get_longest_sheet_name(file_path):
    workbook = load_workbook(file_path, read_only=True)
    sheet_names = workbook.sheetnames
    
    if not sheet_names:
        return None
    
    # Find the sheet with the longest name
    longest_sheet = max(sheet_names, key=len)
    return longest_sheet

# Function to clean column names
def clean_column_names(df):
    """Clean column names by removing leading/trailing spaces and standardizing case"""
    if df is None or df.empty:
        return df
    
    # Create a mapping of old column names to new cleaned names
    column_mapping = {}
    for col in df.columns:
        # Remove leading/trailing spaces and convert to standard case
        cleaned_col = str(col).strip()
        column_mapping[col] = cleaned_col
    
    # Rename columns
    df = df.rename(columns=column_mapping)
    
    # Also clean string data in all columns (remove leading/trailing spaces from cell values)
    for col in df.columns:
        if df[col].dtype == 'object':  # String columns
            df[col] = df[col].astype(str).str.strip()
            # Convert 'nan' strings back to actual NaN
            df[col] = df[col].replace('nan', pd.NA)
    
    return df

# Function to determine file type and read data with caching
def read_data_file_cached(file_path):
    cache_path = get_cache_path(file_path)
    
    # Check if cache is valid
    if is_cache_valid(file_path, cache_path):
        print(f"Loading from cache: {os.path.basename(file_path)}")
        try:
            df = pd.read_pickle(cache_path)
            # Clean column names even for cached data to ensure consistency
            df = clean_column_names(df)
            return df
        except Exception as e:
            print(f"Error loading cache for {os.path.basename(file_path)}: {e}")
            # Fall through to read from source
    
    # Read from source file
    file_extension = os.path.splitext(file_path)[1].lower()
    
    if file_extension == '.xlsx':
        # Handle Excel files
        longest_sheet = get_longest_sheet_name(file_path)
        if longest_sheet:
            print(f"Reading Excel file: {os.path.basename(file_path)}, Sheet: {longest_sheet}")
            df = pd.read_excel(file_path, sheet_name=longest_sheet)
        else:
            print(f"No sheets found in {os.path.basename(file_path)}")
            return None
    elif file_extension == '.csv':
        # Handle CSV files
        print(f"Reading CSV file: {os.path.basename(file_path)}")
        df = pd.read_csv(file_path)
    else:
        print(f"Unsupported file type: {file_extension}")
        return None
    
    # Clean column names and data
    df = clean_column_names(df)
    print(f"Cleaned columns for {os.path.basename(file_path)}: {list(df.columns)}")
    
    # Cache the cleaned data
    try:
        os.makedirs(os.path.dirname(cache_path), exist_ok=True)
        df.to_pickle(cache_path)
        print(f"Cached (cleaned): {os.path.basename(file_path)}")
    except Exception as e:
        print(f"Error caching {os.path.basename(file_path)}: {e}")
    
    return df

# List to store dataframes from each file
all_dfs = []

# Process each file (Excel and CSV) with caching
start_time = time.time()
print("Starting file processing with individual caching...")

for file in all_files:
    try:
        # Read data from the file (with caching)
        df_file = read_data_file_cached(file)
        
        if df_file is not None:
            # Append to our list of dataframes
            all_dfs.append(df_file)
        
    except Exception as e:
        print(f"Error processing {os.path.basename(file)}: {str(e)}")

print(f"File processing finished in {time.time() - start_time:.2f} seconds")

# Now we have all the data in all_dfs list    # This block only runs if we're processing Excel files (no cache was found)



if 'all_dfs' in locals():
    # Combine all dataframes if we have any
    if all_dfs:
        # Clean and standardize all dataframes before combining
        cleaned_dfs = []
        all_columns = set()
        
        # First pass: collect all unique column names and clean dataframes
        for i, df_file in enumerate(all_dfs):
            if df_file is not None and not df_file.empty:
                # Clean the dataframe
                df_cleaned = clean_column_names(df_file)
                cleaned_dfs.append(df_cleaned)
                all_columns.update(df_cleaned.columns)
                print(f"DataFrame {i+1} columns: {list(df_cleaned.columns)}")
        
        print(f"All unique columns found: {sorted(all_columns)}")
        
        # Combine all cleaned dataframes
        combined_df = pd.concat(cleaned_dfs, ignore_index=True, sort=False)
        print(f"Combined dataframe columns: {list(combined_df.columns)}")
        
        # Filter out specific categories
        categories_to_exclude = ['JSP', 'Remove', 'FOC-R', 'EMI', '(blank)', 'PRO', 'WRT', 'Others']
        if 'Category' in combined_df.columns:
            print(f"Filtering out categories: {', '.join(categories_to_exclude)}")
            before_count = len(combined_df)
            combined_df = combined_df[~combined_df['Category'].isin(categories_to_exclude)]
            after_count = len(combined_df)
            print(f"Removed {before_count - after_count} rows with excluded categories")
        
        # Process the date column to extract day, month, and year
        if 'Date' in combined_df.columns:
            print("Processing date column to extract day, month, and year...")
            # Convert to datetime if it's not already
            if not pd.api.types.is_datetime64_dtype(combined_df['Date']):
                combined_df['Date'] = pd.to_datetime(combined_df['Date'], errors='coerce')
            
            # Extract day, month and year
            combined_df['Day'] = combined_df['Date'].dt.day
            combined_df['Month'] = combined_df['Date'].dt.month_name()
            combined_df['Year'] = combined_df['Date'].dt.year
            print("Date columns added")
        
        print("\nCombined data shape:", combined_df.shape)
        print("\nCombined data columns:", combined_df.columns.tolist())
        
        # Display the first few rows of the combined data
        print("\nFirst few rows of combined data:")
        print(combined_df.head())
        
        # Save the combined dataframe to disk for future use
        try:
            # Create cache folder if it doesn't exist
            os.makedirs(os.path.dirname(combined_cache_file), exist_ok=True)
                
            # Save the dataframe
            combined_df.to_pickle(combined_cache_file)
            print(f"\nData saved to {combined_cache_file} for future use")
        except Exception as e:
            print(f"\nError saving data to cache: {str(e)}")
    else:
        print("No data was loaded from the files.")
        combined_df = None

# Load the data from the cache
# Load cached data
if os.path.exists(combined_cache_file):
    df = pd.read_pickle(combined_cache_file)
    print("Data loaded successfully!")
    print(f"Data shape: {df.shape}")
    print(f"Columns: {df.columns.tolist()}")
else:
    print("No cache file found. Please run the script to process files first.")
    df = pd.DataFrame()

# Initialize Dash app
app = dash.Dash(__name__)

# Define metrics for comparison
metrics = ['Items viewed', 'Items added to cart', 'Items purchased', 'Item revenue', 'Sessions']

# Custom CSS styles
app.index_string = '''
<!DOCTYPE html>
<html>
    <head>
        {%metas%}
        <title>{%title%}</title>
        {%favicon%}
        {%css%}
        <style>
            body {
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                margin: 0;
                padding: 20px;
                background-color: #f5f5f5;
            }
            .dashboard-header {
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                padding: 20px;
                border-radius: 10px;
                margin-bottom: 20px;
                text-align: center;
                box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
            }
            .filter-section {
                background: white;
                padding: 20px;
                border-radius: 10px;
                margin-bottom: 20px;
                box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
            }
            .comparison-section {
                background: white;
                padding: 20px;
                border-radius: 10px;
                margin-bottom: 20px;
                box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
            }
            .positive-change {
                color: #28a745;
                font-weight: bold;
            }
            .negative-change {
                color: #dc3545;
                font-weight: bold;
            }            .neutral-change {
                color: #6c757d;
                font-weight: bold;
            }
            .dash-table-container .dash-spreadsheet-container .dash-spreadsheet-inner table {
                border-collapse: separate !important;
                border-spacing: 0 !important;
            }
            .dash-table-container .dash-header {
                border: 1px solid #ddd !important;
            }            .dash-table-container .dash-header.dash-header--merged {
                background-color: #e9ecef !important;
                font-weight: bold !important;
                border-bottom: 2px solid #667eea !important;
            }            .Select-multi-value-wrapper {
                max-height: 60px !important;
                overflow-y: auto !important;
            }
            .Select-value {
                background-color: #667eea !important;
                color: white !important;
                border-color: #667eea !important;
            }
            .Select-value-icon {
                border-right-color: #5a6fd8 !important;
            }
            .Select-value-icon:hover {
                background-color: #5a6fd8 !important;
                color: white !important;
            }
            /* Modern dropdown styling */
            .Select-control {
                border: 1px solid #ddd !important;
                border-radius: 4px !important;
            }
            .Select-control:hover {
                border-color: #667eea !important;
            }            .Select--is-focused .Select-control {
                border-color: #667eea !important;
                box-shadow: 0 0 0 1px #667eea !important;
            }            /* Disabled dropdown styling */
            .Select--is-disabled .Select-control {
                background-color: #f8f9fa !important;
                color: #6c757d !important;
                border-color: #ddd !important;
                cursor: not-allowed !important;
            }            .Select--is-disabled .Select-placeholder {
                color: #6c757d !important;
            }
            /* Hide sort arrows on merged/parent headers only */
            .dash-table-container .dash-header.dash-header--merged .column-header--sort {
                display: none !important;
            }
            .dash-table-container .dash-header.dash-header--merged:hover {
                cursor: default !important;
            }            /* Ensure child column headers show sort arrows */
            .dash-table-container .dash-header:not(.dash-header--merged) .column-header--sort {
                display: block !important;
            }
            /* Styling for exclude dropdowns */
            .exclude-dropdown .Select-control {
                border-color: #dc3545 !important;
                background-color: #fff5f5 !important;
            }
            .exclude-dropdown .Select-control:hover {
                border-color: #c82333 !important;
            }
            .exclude-dropdown .Select--is-focused .Select-control {
                border-color: #dc3545 !important;
                box-shadow: 0 0 0 1px #dc3545 !important;
            }
        </style>
    </head>
    <body>
        {%app_entry%}
        <footer>
            {%config%}
            {%scripts%}
            {%renderer%}
        </footer>
    </body>
</html>
'''

def get_available_months():
    """Get list of available months from the data"""
    if not df.empty and 'Month' in df.columns and 'Year' in df.columns:
        # Create month-year combinations
        df['Month_Year'] = df['Month'].astype(str) + ' ' + df['Year'].astype(str)
        return sorted(df['Month_Year'].unique())
    return []

def get_available_brands():
    """Get list of available brands"""
    if not df.empty and 'Brand' in df.columns:
        try:
            # Convert to string and handle mixed types
            brands = df['Brand'].dropna().astype(str).unique().tolist()
            # Filter out empty strings and sort
            brands = [brand for brand in brands if brand.strip() != '' and brand != 'nan']
            print(f"Available brands found: {len(brands)} - {brands[:10]}...")  # Show first 10 for debugging
            return sorted(brands)
        except Exception as e:
            print(f"Error getting brands: {e}")
            return []
    return []

def get_available_categories():
    """Get list of available categories"""
    if not df.empty and 'Category' in df.columns:
        try:
            # Convert to string and handle mixed types
            categories = df['Category'].dropna().astype(str).unique().tolist()
            # Filter out empty strings and sort
            categories = [cat for cat in categories if cat.strip() != '' and cat != 'nan']
            print(f"Available categories found: {len(categories)} - {categories[:10]}...")  # Show first 10 for debugging
            return sorted(categories)
        except Exception as e:
            print(f"Error getting categories: {e}")
            return []
    return []

def get_available_days():
    """Get list of available days"""
    if not df.empty and 'Day' in df.columns:
        try:
            # Convert to numeric, then to int, then to string
            days = df['Day'].dropna().astype(float).astype(int).unique().tolist()
            days = sorted([day for day in days if 1 <= day <= 31])  # Valid day range
            return [str(day) for day in days]
        except Exception as e:
            print(f"Error getting days: {e}")
            return []
    return []

def filter_data(month1, month2, month1_days, month2_days, brand_filter, category_filter, day_filter):
    """Filter data based on selected criteria"""
    if df.empty:
        print("DataFrame is empty")
        return pd.DataFrame(), pd.DataFrame()
    
    try:
        # Create month-year column for filtering
        df_filtered = df.copy()
        df_filtered['Month_Year'] = df_filtered['Month'].astype(str) + ' ' + df_filtered['Year'].astype(str)
        
        print(f"Total rows before month filtering: {len(df_filtered)}")
        print(f"Available months: {df_filtered['Month_Year'].unique()}")
        print(f"Filtering for months: {month1}, {month2}")
        
        # Filter by months
        df_month1 = df_filtered[df_filtered['Month_Year'] == month1].copy()
        df_month2 = df_filtered[df_filtered['Month_Year'] == month2].copy()
        
        print(f"Rows for {month1}: {len(df_month1)}")
        print(f"Rows for {month2}: {len(df_month2)}")
        
        # Clean filter lists - remove SELECT_ALL from filters if present
        month1_days = [day for day in (month1_days or []) if day != 'SELECT_ALL']
        month2_days = [day for day in (month2_days or []) if day != 'SELECT_ALL']
        brand_filter = [brand for brand in (brand_filter or []) if brand != 'SELECT_ALL']
        category_filter = [cat for cat in (category_filter or []) if cat != 'SELECT_ALL']
        day_filter = [day for day in (day_filter or []) if day != 'SELECT_ALL']
        
        print(f"Applied filters:")
        print(f"  Brand filter: {brand_filter}")
        print(f"  Category filter: {category_filter}")
        print(f"  Day filter: {day_filter}")
        print(f"  Month1 days: {month1_days}")
        print(f"  Month2 days: {month2_days}")
        
        # Handle parent day filters first (month-specific day filtering)
        # Check if parent day filters are active (not empty)
        month1_days_active = month1_days and len(month1_days) > 0
        month2_days_active = month2_days and len(month2_days) > 0
        
        if month1_days_active:
            # Apply day filter for first month
            try:
                day_values = [int(day) for day in month1_days]
                if day_values:
                    print(f"Filtering {month1} for days: {day_values}")
                    df_month1 = df_month1[df_month1['Day'].astype(float).astype(int).isin(day_values)]
                    print(f"Rows after day filter for {month1}: {len(df_month1)}")
            except ValueError as ve:
                print(f"Invalid month1 day filter values: {month1_days}, error: {ve}")
        
        if month2_days_active:
            # Apply day filter for second month
            try:
                day_values = [int(day) for day in month2_days]
                if day_values:
                    print(f"Filtering {month2} for days: {day_values}")
                    df_month2 = df_month2[df_month2['Day'].astype(float).astype(int).isin(day_values)]
                    print(f"Rows after day filter for {month2}: {len(df_month2)}")
            except ValueError as ve:
                print(f"Invalid month2 day filter values: {month2_days}, error: {ve}")
        
        # Only apply child day filter if parent day filters are not active
        if not month1_days_active and not month2_days_active and day_filter and len(day_filter) > 0:
            # Convert Day column to numeric for comparison
            try:
                day_values = [int(day) for day in day_filter]
                if day_values:  # Only filter if there are valid day values
                    print(f"Filtering both months for days: {day_values}")
                    df_month1 = df_month1[df_month1['Day'].astype(float).astype(int).isin(day_values)]
                    df_month2 = df_month2[df_month2['Day'].astype(float).astype(int).isin(day_values)]
                    print(f"Rows after child day filter - {month1}: {len(df_month1)}, {month2}: {len(df_month2)}")
            except ValueError as ve:
                print(f"Invalid day filter values: {day_filter}, error: {ve}")
        
        # Handle multi-select filters for brand and category
        # Brand filter - only apply if not empty
        if brand_filter and len(brand_filter) > 0:
            print(f"Available brands in {month1}: {sorted(df_month1['Brand'].astype(str).unique())}")
            print(f"Available brands in {month2}: {sorted(df_month2['Brand'].astype(str).unique())}")
            print(f"Requested brand filter: {brand_filter}")
            
            # Check if any of the requested brands exist in the data
            brands_in_month1 = set(df_month1['Brand'].astype(str).unique())
            brands_in_month2 = set(df_month2['Brand'].astype(str).unique())
            requested_brands = set(brand_filter)
            
            print(f"Brands found in {month1}: {requested_brands.intersection(brands_in_month1)}")
            print(f"Brands found in {month2}: {requested_brands.intersection(brands_in_month2)}")
            
            # Convert Brand column to string for comparison and apply filter
            df_month1 = df_month1[df_month1['Brand'].astype(str).isin(brand_filter)]
            df_month2 = df_month2[df_month2['Brand'].astype(str).isin(brand_filter)]
            print(f"Rows after brand filter - {month1}: {len(df_month1)}, {month2}: {len(df_month2)}")
            
        # Category filter - only apply if not empty
        if category_filter and len(category_filter) > 0:
            print(f"Available categories in {month1}: {sorted(df_month1['Category'].astype(str).unique())}")
            print(f"Available categories in {month2}: {sorted(df_month2['Category'].astype(str).unique())}")
            print(f"Requested category filter: {category_filter}")
            
            # Check if any of the requested categories exist in the data
            categories_in_month1 = set(df_month1['Category'].astype(str).unique())
            categories_in_month2 = set(df_month2['Category'].astype(str).unique())
            requested_categories = set(category_filter)
            
            print(f"Categories found in {month1}: {requested_categories.intersection(categories_in_month1)}")
            print(f"Categories found in {month2}: {requested_categories.intersection(categories_in_month2)}")
            
            # Convert Category column to string for comparison and apply filter
            df_month1 = df_month1[df_month1['Category'].astype(str).isin(category_filter)]
            df_month2 = df_month2[df_month2['Category'].astype(str).isin(category_filter)]
            print(f"Rows after category filter - {month1}: {len(df_month1)}, {month2}: {len(df_month2)}")
            
            # Additional debugging: Show sample data that matches
            if len(df_month1) > 0:
                print(f"Sample data from {month1} after filtering:")
                print(df_month1[['Brand', 'Category']].head(3).to_string())
            if len(df_month2) > 0:
                print(f"Sample data from {month2} after filtering:")
                print(df_month2[['Brand', 'Category']].head(3).to_string())
        
        print(f"Final results - {month1}: {len(df_month1)} rows, {month2}: {len(df_month2)} rows")
        
        return df_month1, df_month2
        
    except Exception as e:
        print(f"Error filtering data: {e}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame(), pd.DataFrame()

def calculate_percentage_change(val1, val2):
    """Calculate percentage change between two values"""
    if pd.isna(val1) or pd.isna(val2) or val1 == 0:
        return 0
    return ((val2 - val1) / val1) * 100

def format_percentage(value):
    """Format percentage with color coding"""
    if value > 0:
        return f"+{value:.1f}%"
    elif value < 0:
        return f"{value:.1f}%"
    else:
        return "0.0%"

def get_color_style(value):
    """Get color style based on percentage change"""
    if value > 0:
        return {'color': '#28a745', 'fontWeight': 'bold'}  # Green for positive
    elif value < 0:
        return {'color': '#dc3545', 'fontWeight': 'bold'}  # Red for negative
    else:
        return {'color': '#6c757d', 'fontWeight': 'bold'}  # Gray for neutral

# App layout
app.layout = html.Div([
    html.Div([
        html.H1("📊 Monthly Comparison Dashboard", 
                style={'margin': '0', 'fontSize': '2.5rem'}),
        html.P("Compare performance metrics between two months with detailed analytics",
               style={'margin': '10px 0 0 0', 'fontSize': '1.1rem', 'opacity': '0.9'})
    ], className='dashboard-header'),
    
    # Tips section
    html.Div([
        html.Details([
            html.Summary("💡 How to Use This Dashboard", 
                        style={'fontWeight': 'bold', 'fontSize': '16px', 'cursor': 'pointer', 'marginBottom': '10px'}),
            html.Div([
                html.P("📍 **Different Month Comparison**: Select two different months to compare their performance.", style={'margin': '5px 0'}),
                html.P("📍 **Same Month Comparison**: Select the same month twice, then use 'Select Days for First Month' and 'Select Days for Second Month' to create distinct periods (e.g., Week 1 vs Week 2).", style={'margin': '5px 0'}),
                html.P("📍 **Filtering**: Use brand, category, and day filters to focus on specific data. Use 'Select All' to include everything or clear filters.", style={'margin': '5px 0'}),
                html.P("📍 **Empty Results**: If one period has no data, it will show 0 values for comparison. If both periods are empty, you'll see helpful suggestions.", style={'margin': '5px 0'})
            ], style={'backgroundColor': '#f8f9fa', 'padding': '15px', 'borderRadius': '5px', 'marginTop': '10px'})
        ])
    ], style={'marginBottom': '20px', 'padding': '10px', 'border': '1px solid #e0e0e0', 'borderRadius': '8px', 'backgroundColor': '#fafafa'}),
    
    html.Div([
        html.H3("🔧 Filters & Comparison Settings", style={'marginBottom': '20px', 'color': '#333'}),
          html.Div([
            html.Div([
                html.Label("Select First Month:", style={'fontWeight': 'bold', 'marginBottom': '5px', 'display': 'block'}),
                dcc.Dropdown(
                    id='month1-dropdown',
                    options=[{'label': month, 'value': month} for month in get_available_months()],
                    value=get_available_months()[0] if get_available_months() else None,
                    style={'marginBottom': '15px'}
                )
            ], style={'width': '48%', 'display': 'inline-block', 'marginRight': '4%'}),
            
            html.Div([
                html.Label("Select Second Month:", style={'fontWeight': 'bold', 'marginBottom': '5px', 'display': 'block'}),
                dcc.Dropdown(
                    id='month2-dropdown',
                    options=[{'label': month, 'value': month} for month in get_available_months()],
                    value=get_available_months()[1] if len(get_available_months()) > 1 else None,
                    style={'marginBottom': '15px'}
                )
            ], style={'width': '48%', 'display': 'inline-block'})
        ]),
          html.Div([
            html.Div([
                html.Label("Select Days for First Month:", style={'fontWeight': 'bold', 'marginBottom': '5px', 'display': 'block'}),
                dcc.Dropdown(
                    id='month1-days-filter',
                    options=[{'label': '🔲 Select All', 'value': 'SELECT_ALL'}] + [{'label': day, 'value': day} for day in get_available_days()],
                    value=[],
                    multi=True,
                    placeholder="Select days for first month (use Select All to choose all)...",
                    style={'marginBottom': '15px'}
                )
            ], style={'width': '48%', 'display': 'inline-block', 'marginRight': '4%'}),
            
            html.Div([
                html.Label("Select Days for Second Month:", style={'fontWeight': 'bold', 'marginBottom': '5px', 'display': 'block'}),
                dcc.Dropdown(
                    id='month2-days-filter',
                    options=[{'label': '🔲 Select All', 'value': 'SELECT_ALL'}] + [{'label': day, 'value': day} for day in get_available_days()],
                    value=[],
                    multi=True,
                    placeholder="Select days for second month (use Select All to choose all)...",
                    style={'marginBottom': '15px'}
                )
            ], style={'width': '48%', 'display': 'inline-block'})
        ]),html.Div([            html.Div([
                html.Label("Brand Filter:", style={'fontWeight': 'bold', 'marginBottom': '5px', 'display': 'block'}),                dcc.Dropdown(
                    id='brand-filter',
                    options=[{'label': '🔲 Select All', 'value': 'SELECT_ALL'}] + [{'label': brand, 'value': brand} for brand in get_available_brands()],
                    value=[],
                    multi=True,
                    placeholder="Select brands (blank = all, use Select All to choose all)...",
                    style={'marginBottom': '15px'}
                )
            ], style={'width': '32%', 'display': 'inline-block', 'marginRight': '2%'}),
            
            html.Div([
                html.Label("Category Filter:", style={'fontWeight': 'bold', 'marginBottom': '5px', 'display': 'block'}),                dcc.Dropdown(
                    id='category-filter',
                    options=[{'label': '🔲 Select All', 'value': 'SELECT_ALL'}] + [{'label': cat, 'value': cat} for cat in get_available_categories()],
                    value=[],
                    multi=True,
                    placeholder="Select categories (blank = all, use Select All to choose all)...",
                    style={'marginBottom': '15px'}
                )
            ], style={'width': '32%', 'display': 'inline-block', 'marginRight': '2%'}),
            
            html.Div([
                html.Label("Day Filter:", style={'fontWeight': 'bold', 'marginBottom': '5px', 'display': 'block'}),                dcc.Dropdown(
                    id='day-filter',
                    options=[{'label': '🔲 Select All', 'value': 'SELECT_ALL'}] + [{'label': day, 'value': day} for day in get_available_days()],
                    value=[],
                    multi=True,
                    placeholder="Select days (blank = all, use Select All to choose all)...",
                    style={'marginBottom': '15px'}
                )
            ], style={'width': '32%', 'display': 'inline-block'})
        ]),
  
    ], className='filter-section'),
      html.Div([
        html.H3("📊 Category Comparison", 
                style={'marginBottom': '20px', 'color': '#333'}),
        html.Div([
            html.P("📋 Select two months and adjust filters to see comparison data. For same-month comparisons, use the month-specific day filters to create distinct periods.", 
                   style={'textAlign': 'center', 'color': '#666', 'fontSize': '16px', 'padding': '20px'})
        ], id='category-comparison-table')
    ], className='comparison-section'),
    
    html.Div([
        html.H3("🛍️ Item Name Comparison", 
                style={'marginBottom': '20px', 'color': '#333'}),
        html.Div([
            html.P("📋 Select two months and adjust filters to see comparison data. For same-month comparisons, use the month-specific day filters to create distinct periods.", 
                   style={'textAlign': 'center', 'color': '#666', 'fontSize': '16px', 'padding': '20px'})
        ], id='item-comparison-table')
    ], className='comparison-section')
])

@app.callback(
    [Output('category-comparison-table', 'children'),
     Output('item-comparison-table', 'children')],
    [Input('month1-dropdown', 'value'),
     Input('month2-dropdown', 'value'),
     Input('month1-days-filter', 'value'),
     Input('month2-days-filter', 'value'),
     Input('brand-filter', 'value'),
     Input('category-filter', 'value'),
     Input('day-filter', 'value')]
)
def update_comparison_tables(month1, month2, month1_days, month2_days, brand_filter, category_filter, day_filter):
    if not month1 or not month2 or df.empty:
        return html.Div("Please select both months to compare."), html.Div("Please select both months to compare.")
    
    # Handle None values for filters (convert to empty list)
    month1_days = month1_days or []
    month2_days = month2_days or []
    brand_filter = brand_filter or []
    category_filter = category_filter or []
    day_filter = day_filter or []
    
    print(f"\n=== CALLBACK DEBUG ===")
    print(f"Selected filters:")
    print(f"  Month1: {month1}")
    print(f"  Month2: {month2}")
    if month1 == month2:
        print(f"  🔄 SAME MONTH COMPARISON DETECTED")
    print(f"  Brand filter: {brand_filter}")
    print(f"  Category filter: {category_filter}")
    print(f"  Day filter: {day_filter}")
    print(f"  Month1 days: {month1_days}")
    print(f"  Month2 days: {month2_days}")
    
    # Filter data
    df_month1, df_month2 = filter_data(month1, month2, month1_days, month2_days, brand_filter, category_filter, day_filter)
    
    # Check if BOTH months are empty (only return error if no data exists at all)
    if df_month1.empty and df_month2.empty:
        # Better error handling - check what filters are actually applied
        has_brand_filter = brand_filter and len([b for b in brand_filter if b != 'SELECT_ALL']) > 0
        has_category_filter = category_filter and len([c for c in category_filter if c != 'SELECT_ALL']) > 0
        has_day_filter = day_filter and len([d for d in day_filter if d != 'SELECT_ALL']) > 0
        has_month1_days = month1_days and len([d for d in month1_days if d != 'SELECT_ALL']) > 0
        has_month2_days = month2_days and len([d for d in month2_days if d != 'SELECT_ALL']) > 0
        
        error_msg = "❌ No data available for the selected criteria in either month."
        
        if month1 == month2:
            # Same month comparison - need different criteria to distinguish periods
            if not has_month1_days and not has_month2_days and not has_day_filter:
                error_msg += " \n\n💡 **For same month comparisons:** Use the 'Select Days for First Month' and 'Select Days for Second Month' filters to create distinct time periods (e.g., first half vs second half of the month)."
            elif has_month1_days or has_month2_days:
                available_days = sorted([str(d) for d in df[df['Month_Year'] == month1]['Day'].unique() if not pd.isna(d)])
                if available_days:
                    error_msg += f" \n\n📅 **Available days in {month1}:** {', '.join(available_days)}. Please select days that exist in the data."
                else:
                    error_msg += f" \n\n⚠️ No valid days found in {month1}. Please check your data."
            else:
                error_msg += " \n\n🔍 **Try these solutions:**\n• Remove some filters to broaden your selection\n• Check if the selected brands/categories exist in this month\n• Verify the day ranges are valid"
        else:
            # Different months - check specific filter issues
            filter_info = []
            if has_brand_filter:
                filter_info.append(f"Brand: {[b for b in brand_filter if b != 'SELECT_ALL']}")
            if has_category_filter:
                filter_info.append(f"Category: {[c for c in category_filter if c != 'SELECT_ALL']}")
            if has_day_filter:
                filter_info.append(f"Days: {[d for d in day_filter if d != 'SELECT_ALL']}")
            
            if filter_info:
                error_msg += f" \n\n🔧 **Applied filters:** {', '.join(filter_info)}.\n\n💡 **Try these solutions:**\n• Use 'Select All' to remove restrictive filters\n• Check if the selected items exist in both months\n• Consider broadening your selection criteria"
            else:
                error_msg += f" \n\n⚠️ No data found in either {month1} or {month2}. Please verify your data contains these months."
        
        print(f"ERROR: {error_msg}")
        # Create a nicer error display
        error_display = html.Div([
            html.H4("🚫 No Data Found", style={'color': '#d32f2f', 'marginBottom': '10px'}),
            dcc.Markdown(error_msg, style={'whiteSpace': 'pre-line', 'lineHeight': '1.6'})
        ], style={
            'padding': '20px', 
            'border': '2px solid #ffcdd2', 
            'borderRadius': '8px', 
            'backgroundColor': '#ffebee',
            'margin': '20px 0'
        })
        return error_display, error_display
    
    # If we reach here, at least one month has data, which is what we want
    # The comparison functions will handle empty dataframes properly by showing 0 values
    print(f"✅ Proceeding with comparison:")
    print(f"   {month1}: {len(df_month1)} rows")
    print(f"   {month2}: {len(df_month2)} rows")
    print(f"   Note: Empty months will show 0 values in comparison tables")
    
    # Category Comparison
    category_comparison = create_category_comparison(df_month1, df_month2, month1, month2)
    
    # Item Comparison
    item_comparison = create_item_comparison(df_month1, df_month2, month1, month2)
    
    return category_comparison, item_comparison

def create_category_comparison(df_month1, df_month2, month1, month2):
    """Create category comparison table"""
    try:
        # Ensure Category column is string type and clean data
        df_month1 = df_month1.copy()
        df_month2 = df_month2.copy()
        
        # Handle empty dataframes - create empty aggregated dataframes if needed
        if not df_month1.empty:
            df_month1['Category'] = df_month1['Category'].astype(str)
            df_month1 = df_month1[df_month1['Category'] != 'nan']
            agg_month1 = df_month1.groupby('Category')[metrics].sum().reset_index()
        else:
            # Create empty dataframe with proper structure
            agg_month1 = pd.DataFrame(columns=['Category'] + metrics)
            
        if not df_month2.empty:
            df_month2['Category'] = df_month2['Category'].astype(str)
            df_month2 = df_month2[df_month2['Category'] != 'nan']
            agg_month2 = df_month2.groupby('Category')[metrics].sum().reset_index()
        else:
            # Create empty dataframe with proper structure
            agg_month2 = pd.DataFrame(columns=['Category'] + metrics)
        
        # Handle same month comparison by creating unique suffixes
        if month1 == month2:
            suffix1 = f'{month1}_Period1'
            suffix2 = f'{month2}_Period2'
        else:
            suffix1 = f'_{month1}'
            suffix2 = f'_{month2}'
        
        # Merge the data with proper suffixes
        # Using 'outer' join to include ALL categories from both periods
        # Categories appearing in only one period will show 0 for the missing period
        comparison = pd.merge(agg_month1, agg_month2, on='Category', suffixes=(suffix1, suffix2), how='outer').fillna(0)
        
        # If both dataframes were empty, we shouldn't reach here, but just in case
        if comparison.empty:
            return html.Div("No category data available for comparison.")
        
        # Calculate percentage changes
        for metric in metrics:
            col1 = f'{metric}{suffix1}'
            col2 = f'{metric}{suffix2}'
            comparison[f'{metric}_change'] = comparison.apply(
                lambda row: calculate_percentage_change(row[col1], row[col2]), axis=1
            )        # Create table data
        table_data = []
        for _, row in comparison.iterrows():
            row_data = {'Category': row['Category']}
            
            for metric in metrics:
                col1 = f'{metric}{suffix1}'
                col2 = f'{metric}{suffix2}'
                change_col = f'{metric}_change'
                
                # Store actual numeric values for sorting, but format for display
                row_data[f'{month1 if month1 != month2 else "Period1"}_{metric}'] = row[col1]
                row_data[f'{month2 if month1 != month2 else "Period2"}_{metric}'] = row[col2]
                row_data[f'Change_{metric}'] = row[change_col] / 100  # Convert to decimal for percentage formatting
            
            table_data.append(row_data)
        
        # Create multi-level columns with proper hierarchy
        columns = [
            {'name': ['', 'Category'], 'id': 'Category', 'type': 'text'}
        ]
        
        for metric in metrics:
            period1_label = month1 if month1 != month2 else "Period 1"
            period2_label = month2 if month1 != month2 else "Period 2"
            columns.extend([
                {'name': [metric, period1_label], 'id': f'{month1 if month1 != month2 else "Period1"}_{metric}', 'type': 'numeric', 'format': {'specifier': ',.0f'}},
                {'name': [metric, period2_label], 'id': f'{month2 if month1 != month2 else "Period2"}_{metric}', 'type': 'numeric', 'format': {'specifier': ',.0f'}},
                {'name': [metric, '% Change'], 'id': f'Change_{metric}', 'type': 'numeric', 'format': {'specifier': '+.1%'}}
            ])          # Style data conditionally for percentage columns
        style_data_conditional = []
        for metric in metrics:
            change_col = f'Change_{metric}'
            # Apply styling based on cell value, not row index
            style_data_conditional.extend([
                {
                    'if': {
                        'filter_query': f'{{{change_col}}} > 0',
                        'column_id': change_col
                    },
                    'color': '#28a745',
                    'fontWeight': 'bold'
                },
                {
                    'if': {
                        'filter_query': f'{{{change_col}}} < 0',
                        'column_id': change_col
                    },
                    'color': '#dc3545',
                    'fontWeight': 'bold'
                },
                {
                    'if': {
                        'filter_query': f'{{{change_col}}} = 0',
                        'column_id': change_col
                    },
                    'color': '#6c757d',
                    'fontWeight': 'bold'
                }
            ])
        return dash_table.DataTable(
            data=table_data,
            columns=columns,
            merge_duplicate_headers=True,
            style_table={'overflowX': 'auto'},
            style_cell={
                'textAlign': 'left',
                'padding': '12px',
                'fontFamily': 'Arial, sans-serif',
                'fontSize': '14px',
                'border': '1px solid #ddd'
            },
            style_header={
                'backgroundColor': '#f8f9fa',
                'fontWeight': 'bold',
                'color': '#333',
                'border': '1px solid #ddd',
                'textAlign': 'center'
            },
            style_data_conditional=style_data_conditional,
            page_size=20,
            sort_action='native',
            sort_mode='multi',
            filter_action='native'
        )
    
    except Exception as e:
        print(f"Error creating category comparison: {e}")
        return html.Div(f"Error creating category comparison table: {str(e)}")

def create_item_comparison(df_month1, df_month2, month1, month2):
    """Create item comparison table"""
    try:
        # Ensure Item name column is string type and clean data
        df_month1 = df_month1.copy()
        df_month2 = df_month2.copy()
        
        # Handle empty dataframes - create empty aggregated dataframes if needed
        if not df_month1.empty:
            df_month1['Item name'] = df_month1['Item name'].astype(str)
            df_month1 = df_month1[df_month1['Item name'] != 'nan']
            agg_month1 = df_month1.groupby('Item name')[metrics].sum().reset_index()
        else:
            # Create empty dataframe with proper structure
            agg_month1 = pd.DataFrame(columns=['Item name'] + metrics)
            
        if not df_month2.empty:
            df_month2['Item name'] = df_month2['Item name'].astype(str)
            df_month2 = df_month2[df_month2['Item name'] != 'nan']
            agg_month2 = df_month2.groupby('Item name')[metrics].sum().reset_index()
        else:
            # Create empty dataframe with proper structure
            agg_month2 = pd.DataFrame(columns=['Item name'] + metrics)
        
        # Handle same month comparison by creating unique suffixes
        if month1 == month2:
            suffix1 = f'{month1}_Period1'
            suffix2 = f'{month2}_Period2'
        else:
            suffix1 = f'_{month1}'
            suffix2 = f'_{month2}'
        
        # Merge the data with proper suffixes
        # Using 'outer' join to include ALL items from both periods
        # Items appearing in only one period will show 0 for the missing period
        comparison = pd.merge(agg_month1, agg_month2, on='Item name', suffixes=(suffix1, suffix2), how='outer').fillna(0)
        
        # If both dataframes were empty, we shouldn't reach here, but just in case
        if comparison.empty:
            return html.Div("No item data available for comparison.")
        
        # Calculate percentage changes
        for metric in metrics:
            col1 = f'{metric}{suffix1}'
            col2 = f'{metric}{suffix2}'
            comparison[f'{metric}_change'] = comparison.apply(
                lambda row: calculate_percentage_change(row[col1], row[col2]), axis=1
            )
        
        # Sort by total revenue change (descending) to prioritize most impactful items
        if 'Item revenue_change' in comparison.columns:
            comparison = comparison.sort_values('Item revenue_change', ascending=False)
        
        # Show all items (no limit) - user requested to see complete data
        # comparison = comparison.head(50)  # Commented out to show all items        # Create table data
        table_data = []
        for _, row in comparison.iterrows():
            row_data = {'Item name': row['Item name']}
            
            for metric in metrics:
                col1 = f'{metric}{suffix1}'
                col2 = f'{metric}{suffix2}'
                change_col = f'{metric}_change'
                
                # Store actual numeric values for sorting, but format for display
                row_data[f'{month1 if month1 != month2 else "Period1"}_{metric}'] = row[col1]
                row_data[f'{month2 if month1 != month2 else "Period2"}_{metric}'] = row[col2]
                row_data[f'Change_{metric}'] = row[change_col] / 100  # Convert to decimal for percentage formatting
            
            table_data.append(row_data)
          
        # Create multi-level columns with proper hierarchy
        columns = [
            {'name': ['', 'Item Name'], 'id': 'Item name', 'type': 'text'}
        ]
        
        for metric in metrics:
            period1_label = month1 if month1 != month2 else "Period 1"
            period2_label = month2 if month1 != month2 else "Period 2"
            columns.extend([
                {'name': [metric, period1_label], 'id': f'{month1 if month1 != month2 else "Period1"}_{metric}', 'type': 'numeric', 'format': {'specifier': ',.0f'}},
                {'name': [metric, period2_label], 'id': f'{month2 if month1 != month2 else "Period2"}_{metric}', 'type': 'numeric', 'format': {'specifier': ',.0f'}},
                {'name': [metric, '% Change'], 'id': f'Change_{metric}', 'type': 'numeric', 'format': {'specifier': '+.1%'}}
            ])        
        # Style data conditionally for percentage columns (item comparison)
        style_data_conditional = []
        for metric in metrics:
            change_col = f'Change_{metric}'
            # Apply styling based on cell value, not row index
            style_data_conditional.extend([
                {
                    'if': {
                        'filter_query': f'{{{change_col}}} > 0',
                        'column_id': change_col
                    },
                    'color': '#28a745',
                    'fontWeight': 'bold'
                },
                {
                    'if': {
                        'filter_query': f'{{{change_col}}} < 0',
                        'column_id': change_col
                    },
                    'color': '#dc3545',
                    'fontWeight': 'bold'
                },
                {
                    'if': {
                        'filter_query': f'{{{change_col}}} = 0',
                        'column_id': change_col
                    },
                    'color': '#6c757d',
                    'fontWeight': 'bold'
                }
            ])
        return dash_table.DataTable(
            data=table_data,
            columns=columns,
            merge_duplicate_headers=True,
            style_table={'overflowX': 'auto'},
            style_cell={
                'textAlign': 'left',
                'padding': '12px',
                'fontFamily': 'Arial, sans-serif',
                'fontSize': '14px',
                'border': '1px solid #ddd',
                'maxWidth': '200px',
                'overflow': 'hidden',
                'textOverflow': 'ellipsis'
            },
            style_header={
                'backgroundColor': '#f8f9fa',
                'fontWeight': 'bold',
                'color': '#333',
                'border': '1px solid #ddd',
                'textAlign': 'center'
            },            style_data_conditional=style_data_conditional,
            page_size=20,
            sort_action='native',
            sort_mode='multi',
            filter_action='native',
            tooltip_data=[
                {
                    'Item name': {'value': str(row['Item name']), 'type': 'markdown'}
                } for row in table_data
            ],
            tooltip_duration=None
        )
    
    except Exception as e:
        print(f"Error creating item comparison: {e}")
        return html.Div(f"Error creating item comparison table: {str(e)}")

# Callbacks for Select All functionality
@app.callback(
    Output('brand-filter', 'value'),
    [Input('brand-filter', 'value')],
    [State('brand-filter', 'options')]
)
def handle_brand_select_all(selected_values, options):
    if not selected_values:
        return []
    
    if 'SELECT_ALL' in selected_values:
        # Get all brand values (excluding SELECT_ALL)
        all_brands = [opt['value'] for opt in options if opt['value'] != 'SELECT_ALL']
        return all_brands
    
    return selected_values

@app.callback(
    Output('category-filter', 'value'),
    [Input('category-filter', 'value')],
    [State('category-filter', 'options')]
)
def handle_category_select_all(selected_values, options):
    if not selected_values:
        return []
    
    if 'SELECT_ALL' in selected_values:
        # Get all category values (excluding SELECT_ALL)
        all_categories = [opt['value'] for opt in options if opt['value'] != 'SELECT_ALL']
        return all_categories
    
    return selected_values

@app.callback(
    Output('day-filter', 'value'),
    [Input('day-filter', 'value')],
    [State('day-filter', 'options')]
)
def handle_day_select_all(selected_values, options):
    if not selected_values:
        return []
    
    if 'SELECT_ALL' in selected_values:
        # Get all day values (excluding SELECT_ALL)
        all_days = [opt['value'] for opt in options if opt['value'] != 'SELECT_ALL']
        return all_days
    
    return selected_values

@app.callback(
    Output('month1-days-filter', 'value'),
    [Input('month1-days-filter', 'value')],
    [State('month1-days-filter', 'options')]
)
def handle_month1_days_select_all(selected_values, options):
    if not selected_values:
        return []
    
    if 'SELECT_ALL' in selected_values:
        # Get all day values (excluding SELECT_ALL)
        all_days = [opt['value'] for opt in options if opt['value'] != 'SELECT_ALL']
        return all_days
    
    return selected_values

@app.callback(
    Output('month2-days-filter', 'value'),
    [Input('month2-days-filter', 'value')],
    [State('month2-days-filter', 'options')]
)
def handle_month2_days_select_all(selected_values, options):
    if not selected_values:
        return []
    
    if 'SELECT_ALL' in selected_values:
        # Get all day values (excluding SELECT_ALL)
        all_days = [opt['value'] for opt in options if opt['value'] != 'SELECT_ALL']
        return all_days
    
    return selected_values
@app.callback(
    [Output('day-filter', 'disabled'),
     Output('day-filter', 'placeholder')],
    [Input('month1-days-filter', 'value'),
     Input('month2-days-filter', 'value')]
)
def disable_child_day_filter(month1_days, month2_days):
    """Disable child day filter when parent day filters are active"""
    # Check if either parent day filter has selections
    month1_has_selection = month1_days and len(month1_days) > 0
    month2_has_selection = month2_days and len(month2_days) > 0
    
    if month1_has_selection or month2_has_selection:
        # Disable child day filter and update placeholder
        return True, "Disabled when month-specific day filters are active"
    else:
        # Enable child day filter with normal placeholder
        return False, "Select days (blank = all, use Select All to choose all)..."

if __name__ == '__main__':
    print("\n" + "="*50)
    print("🚀 Starting Comparison Dashboard...")
    print("="*50)
    
    if df.empty:
        print("❌ No data loaded. Please run mainDashboard.py first to generate cache.")
    else:
        print(f"✅ Data loaded successfully: {df.shape[0]} rows, {df.shape[1]} columns")
        print(f"📅 Available months: {len(get_available_months())}")
        print(f"🏢 Available brands: {len(get_available_brands())}")
        print(f"📦 Available categories: {len(get_available_categories())}")
        
    print("\n🌐 Dashboard will be available at: http://127.0.0.I1:8050")
    print("📊 Features:")
    print("   • Month-to-month comparison")
    print("   • Percentage change calculations with color coding")
    print("   • Filtering by Brand, Category, and Day")
    print("   • Category-wise and Item-wise comparison tables")
    print("   • Sortable and filterable tables")
    print("\n" + "="*50)
    
    app.run(debug=False, host='127.0.0.1', port=8050)
