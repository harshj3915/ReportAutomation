import json
import os
import pandas as pd
import dash
from dash import dcc, html, dash_table, callback
from dash.dependencies import Input, Output, State
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import numpy as np

import os
import pandas as pd
import glob
from openpyxl import load_workbook
import time

# Define paths for saved data
cache_folder = os.path.join(os.path.dirname(__file__), 'cache')
cache_file = os.path.join(cache_folder, 'combined_data.pkl')

# Check if cached data exists first, before processing any Excel files
if os.path.exists(cache_file):
    print(f"\nCache file found at {cache_file}")
    print(f"Loading cached data instead of processing Excel files...")
    combined_df = pd.read_pickle(cache_file)
    print("Cached data loaded successfully!")
    
    # Filter out specific categories
    categories_to_exclude = ['JSP', 'Remove', 'FOC-R', 'EMI', '(blank)', 'PRO', 'WRT', 'Others']
    if 'Category' in combined_df.columns:
        print(f"Filtering out categories: {', '.join(categories_to_exclude)}")
        before_count = len(combined_df)
        combined_df = combined_df[~combined_df['Category'].isin(categories_to_exclude)]
        after_count = len(combined_df)
        print(f"Removed {before_count - after_count} rows with excluded categories")
    
    # Process the date column to extract day, month, and year
    if 'Date' in combined_df.columns and not all(col in combined_df.columns for col in ['Day', 'Month', 'Year']):
        print("Processing date column to extract day, month, and year...")
        # Convert to datetime if it's not already
        if not pd.api.types.is_datetime64_dtype(combined_df['Date']):
            combined_df['Date'] = pd.to_datetime(combined_df['Date'], errors='coerce')
        
        # Extract day, month and year
        combined_df['Day'] = combined_df['Date'].dt.day
        combined_df['Month'] = combined_df['Date'].dt.month_name()
        combined_df['Year'] = combined_df['Date'].dt.year
        
        # Save the updated dataframe back to cache
        combined_df.to_pickle(cache_file)
        print("Date columns added and cache updated")
    
    print("\nCombined data shape:", combined_df.shape)
    print("\nCombined data columns:", combined_df.columns.tolist())
    
    # Display the first few rows of the combined data
    print("\nFirst few rows of combined data:")
    print(combined_df.head())
    
else:
    print("No cache found. Processing Excel files...")
    
    try:
        with open('config.json', 'r') as f:
            config_data = json.load(f)
        folder_path = config_data['paths']['MainDashboardData']
    except Exception as e:
        print(f"❌ Error loading config.json: {e}")
        print("Using default DSR folder path...")
        folder_path = "C:\\Users\\91843\\Documents\\VsCode Codes\\ReportAutomation\\test\\DSR"
    print(f"Folder path: {folder_path}")
    
    # Get a list of all Excel files in the folder
    excel_files = glob.glob(os.path.join(folder_path, '*.xlsx'))
    print(f"Found {len(excel_files)} Excel files")
    
    # Function to get the sheet with the longest name from an Excel file
    def get_longest_sheet_name(file_path):
        workbook = load_workbook(file_path, read_only=True)
        sheet_names = workbook.sheetnames
        
        if not sheet_names:
            return None
        
        # Find the sheet with the longest name
        longest_sheet = max(sheet_names, key=len)
        return longest_sheet
    
    # List to store dataframes from each Excel file
    all_dfs = []
    
    # Process each Excel file
    start_time = time.time()
    print("Starting Excel file processing...")
    
    for file in excel_files:
        try:
            # Get the sheet with the longest name
            longest_sheet = get_longest_sheet_name(file)
            
            if longest_sheet:
                print(f"Reading file: {os.path.basename(file)}, Sheet: {longest_sheet}")
                
                # Read the data from the sheet with the longest name
                df = pd.read_excel(file, sheet_name=longest_sheet)
                               
                # Append to our list of dataframes
                all_dfs.append(df)
            else:
                print(f"No sheets found in {os.path.basename(file)}")
                
        except Exception as e:
            print(f"Error processing {os.path.basename(file)}: {str(e)}")
    
    print(f"Excel processing finished in {time.time() - start_time:.2f} seconds")
    
    # Now we have all the data in all_dfs list    # This block only runs if we're processing Excel files (no cache was found)
if 'all_dfs' in locals():
    # Combine all dataframes if we have any
    if all_dfs:
        combined_df = pd.concat(all_dfs, ignore_index=True)
        
        # Filter out specific categories
        categories_to_exclude = ['JSP', 'Remove', 'FOC-R', 'EMI', '(blank)', 'PRO', 'WRT', 'Others']
        if 'Category' in combined_df.columns:
            print(f"Filtering out categories: {', '.join(categories_to_exclude)}")
            before_count = len(combined_df)
            combined_df = combined_df[~combined_df['Category'].isin(categories_to_exclude)]
            after_count = len(combined_df)
            print(f"Removed {before_count - after_count} rows with excluded categories")
        
        # Process the date column to extract day, month, and year
        if 'Date' in combined_df.columns:
            print("Processing date column to extract day, month, and year...")
            # Convert to datetime if it's not already
            if not pd.api.types.is_datetime64_dtype(combined_df['Date']):
                combined_df['Date'] = pd.to_datetime(combined_df['Date'], errors='coerce')
            
            # Extract day, month and year
            combined_df['Day'] = combined_df['Date'].dt.day
            combined_df['Month'] = combined_df['Date'].dt.month_name()
            combined_df['Year'] = combined_df['Date'].dt.year
            print("Date columns added")
        
        print("\nCombined data shape:", combined_df.shape)
        print("\nCombined data columns:", combined_df.columns.tolist())
        
        # Display the first few rows of the combined data
        print("\nFirst few rows of combined data:")
        print(combined_df.head())
        
        # Save the combined dataframe to disk for future use
        try:
            # Create cache folder if it doesn't exist
            if not os.path.exists(cache_folder):
                os.makedirs(cache_folder)
                
            # Save the dataframe
            combined_df.to_pickle(cache_file)
            print(f"\nData saved to {cache_file} for future use")
        except Exception as e:
            print(f"\nError saving data to cache: {str(e)}")
    else:
        print("No data was loaded from the Excel files.")
        combined_df = None

# Load the data from the cache
cache_folder = os.path.join(os.path.dirname(__file__), 'cache')
cache_file = os.path.join(cache_folder, 'combined_data.pkl')

# Load cached data
if os.path.exists(cache_file):
    df = pd.read_pickle(cache_file)
    print("Data loaded successfully!")
    print(f"Data shape: {df.shape}")
    print(f"Columns: {df.columns.tolist()}")
else:
    print("No cache file found. Please run mainDashboard.py first.")
    df = pd.DataFrame()

# Initialize Dash app
app = dash.Dash(__name__)

# Define metrics for comparison
metrics = ['Items viewed', 'Items added to cart', 'Items purchased', 'Item revenue', 'Sessions']

# Custom CSS styles
app.index_string = '''
<!DOCTYPE html>
<html>
    <head>
        {%metas%}
        <title>{%title%}</title>
        {%favicon%}
        {%css%}
        <style>
            body {
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                margin: 0;
                padding: 20px;
                background-color: #f5f5f5;
            }
            .dashboard-header {
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: white;
                padding: 20px;
                border-radius: 10px;
                margin-bottom: 20px;
                text-align: center;
                box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
            }
            .filter-section {
                background: white;
                padding: 20px;
                border-radius: 10px;
                margin-bottom: 20px;
                box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
            }
            .comparison-section {
                background: white;
                padding: 20px;
                border-radius: 10px;
                margin-bottom: 20px;
                box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
            }
            .positive-change {
                color: #28a745;
                font-weight: bold;
            }
            .negative-change {
                color: #dc3545;
                font-weight: bold;
            }            .neutral-change {
                color: #6c757d;
                font-weight: bold;
            }
            .dash-table-container .dash-spreadsheet-container .dash-spreadsheet-inner table {
                border-collapse: separate !important;
                border-spacing: 0 !important;
            }
            .dash-table-container .dash-header {
                border: 1px solid #ddd !important;
            }            .dash-table-container .dash-header.dash-header--merged {
                background-color: #e9ecef !important;
                font-weight: bold !important;
                border-bottom: 2px solid #667eea !important;
            }            .Select-multi-value-wrapper {
                max-height: 60px !important;
                overflow-y: auto !important;
            }
            .Select-value {
                background-color: #667eea !important;
                color: white !important;
                border-color: #667eea !important;
            }
            .Select-value-icon {
                border-right-color: #5a6fd8 !important;
            }
            .Select-value-icon:hover {
                background-color: #5a6fd8 !important;
                color: white !important;
            }
            /* Modern dropdown styling */
            .Select-control {
                border: 1px solid #ddd !important;
                border-radius: 4px !important;
            }
            .Select-control:hover {
                border-color: #667eea !important;
            }            .Select--is-focused .Select-control {
                border-color: #667eea !important;
                box-shadow: 0 0 0 1px #667eea !important;
            }            /* Disabled dropdown styling */
            .Select--is-disabled .Select-control {
                background-color: #f8f9fa !important;
                color: #6c757d !important;
                border-color: #ddd !important;
                cursor: not-allowed !important;
            }            .Select--is-disabled .Select-placeholder {
                color: #6c757d !important;
            }
            /* Hide sort arrows on merged/parent headers only */
            .dash-table-container .dash-header.dash-header--merged .column-header--sort {
                display: none !important;
            }
            .dash-table-container .dash-header.dash-header--merged:hover {
                cursor: default !important;
            }
            /* Ensure child column headers show sort arrows */
            .dash-table-container .dash-header:not(.dash-header--merged) .column-header--sort {
                display: block !important;
            }
        </style>
    </head>
    <body>
        {%app_entry%}
        <footer>
            {%config%}
            {%scripts%}
            {%renderer%}
        </footer>
    </body>
</html>
'''

def get_available_months():
    """Get list of available months from the data"""
    if not df.empty and 'Month' in df.columns and 'Year' in df.columns:
        # Create month-year combinations
        df['Month_Year'] = df['Month'].astype(str) + ' ' + df['Year'].astype(str)
        return sorted(df['Month_Year'].unique())
    return []

def get_available_brands():
    """Get list of available brands"""
    if not df.empty and 'Brand' in df.columns:
        try:
            # Convert to string and handle mixed types
            brands = df['Brand'].dropna().astype(str).unique().tolist()
            # Filter out empty strings and sort
            brands = [brand for brand in brands if brand.strip() != '' and brand != 'nan']
            return sorted(brands)
        except Exception as e:
            print(f"Error getting brands: {e}")
            return []
    return []

def get_available_categories():
    """Get list of available categories"""
    if not df.empty and 'Category' in df.columns:
        try:
            # Convert to string and handle mixed types
            categories = df['Category'].dropna().astype(str).unique().tolist()
            # Filter out empty strings and sort
            categories = [cat for cat in categories if cat.strip() != '' and cat != 'nan']
            return sorted(categories)
        except Exception as e:
            print(f"Error getting categories: {e}")
            return []
    return []

def get_available_days():
    """Get list of available days"""
    if not df.empty and 'Day' in df.columns:
        try:
            # Convert to numeric, then to int, then to string
            days = df['Day'].dropna().astype(float).astype(int).unique().tolist()
            days = sorted([day for day in days if 1 <= day <= 31])  # Valid day range
            return [str(day) for day in days]
        except Exception as e:
            print(f"Error getting days: {e}")
            return []
    return []

def filter_data(month1, month2, month1_days, month2_days, brand_filter, category_filter, day_filter):
    """Filter data based on selected criteria"""
    if df.empty:
        return pd.DataFrame(), pd.DataFrame()
    
    try:
        # Create month-year column for filtering
        df_filtered = df.copy()
        df_filtered['Month_Year'] = df_filtered['Month'].astype(str) + ' ' + df_filtered['Year'].astype(str)
        
        # Filter by months
        df_month1 = df_filtered[df_filtered['Month_Year'] == month1].copy()
        df_month2 = df_filtered[df_filtered['Month_Year'] == month2].copy()
        
        # Handle parent day filters first (month-specific day filtering)
        # Check if parent day filters are active (not empty)
        month1_days_active = month1_days and len(month1_days) > 0
        month2_days_active = month2_days and len(month2_days) > 0
        
        if month1_days_active:
            # Apply day filter for first month
            try:
                day_values = [int(day) for day in month1_days]
                if day_values:
                    df_month1 = df_month1[df_month1['Day'].astype(float).astype(int).isin(day_values)]
            except ValueError as ve:
                print(f"Invalid month1 day filter values: {month1_days}, error: {ve}")
        
        if month2_days_active:
            # Apply day filter for second month
            try:
                day_values = [int(day) for day in month2_days]
                if day_values:
                    df_month2 = df_month2[df_month2['Day'].astype(float).astype(int).isin(day_values)]
            except ValueError as ve:
                print(f"Invalid month2 day filter values: {month2_days}, error: {ve}")
        
        # Only apply child day filter if parent day filters are not active
        if not month1_days_active and not month2_days_active and day_filter and len(day_filter) > 0:
            # Convert Day column to numeric for comparison
            try:
                day_values = [int(day) for day in day_filter]
                if day_values:  # Only filter if there are valid day values
                    df_month1 = df_month1[df_month1['Day'].astype(float).astype(int).isin(day_values)]
                    df_month2 = df_month2[df_month2['Day'].astype(float).astype(int).isin(day_values)]
            except ValueError as ve:
                print(f"Invalid day filter values: {day_filter}, error: {ve}")
        
        # Handle multi-select filters for brand and category
        # Brand filter - only apply if not empty
        if brand_filter and len(brand_filter) > 0:
            # Convert Brand column to string for comparison
            df_month1 = df_month1[df_month1['Brand'].astype(str).isin(brand_filter)]
            df_month2 = df_month2[df_month2['Brand'].astype(str).isin(brand_filter)]
            
        # Category filter - only apply if not empty
        if category_filter and len(category_filter) > 0:
            # Convert Category column to string for comparison
            df_month1 = df_month1[df_month1['Category'].astype(str).isin(category_filter)]
            df_month2 = df_month2[df_month2['Category'].astype(str).isin(category_filter)]
        
        return df_month1, df_month2
        
    except Exception as e:
        print(f"Error filtering data: {e}")
        return pd.DataFrame(), pd.DataFrame()

def calculate_percentage_change(val1, val2):
    """Calculate percentage change between two values"""
    if pd.isna(val1) or pd.isna(val2) or val1 == 0:
        return 0
    return ((val2 - val1) / val1) * 100

def format_percentage(value):
    """Format percentage with color coding"""
    if value > 0:
        return f"+{value:.1f}%"
    elif value < 0:
        return f"{value:.1f}%"
    else:
        return "0.0%"

def get_color_style(value):
    """Get color style based on percentage change"""
    if value > 0:
        return {'color': '#28a745', 'fontWeight': 'bold'}  # Green for positive
    elif value < 0:
        return {'color': '#dc3545', 'fontWeight': 'bold'}  # Red for negative
    else:
        return {'color': '#6c757d', 'fontWeight': 'bold'}  # Gray for neutral

# App layout
app.layout = html.Div([
    html.Div([
        html.H1("📊 Monthly Comparison Dashboard", 
                style={'margin': '0', 'fontSize': '2.5rem'}),
        html.P("Compare performance metrics between two months with detailed analytics",
               style={'margin': '10px 0 0 0', 'fontSize': '1.1rem', 'opacity': '0.9'})
    ], className='dashboard-header'),
    
    html.Div([
        html.H3("🔧 Filters & Comparison Settings", style={'marginBottom': '20px', 'color': '#333'}),
          html.Div([
            html.Div([
                html.Label("Select First Month:", style={'fontWeight': 'bold', 'marginBottom': '5px', 'display': 'block'}),
                dcc.Dropdown(
                    id='month1-dropdown',
                    options=[{'label': month, 'value': month} for month in get_available_months()],
                    value=get_available_months()[0] if get_available_months() else None,
                    style={'marginBottom': '15px'}
                )
            ], style={'width': '48%', 'display': 'inline-block', 'marginRight': '4%'}),
            
            html.Div([
                html.Label("Select Second Month:", style={'fontWeight': 'bold', 'marginBottom': '5px', 'display': 'block'}),
                dcc.Dropdown(
                    id='month2-dropdown',
                    options=[{'label': month, 'value': month} for month in get_available_months()],
                    value=get_available_months()[1] if len(get_available_months()) > 1 else None,
                    style={'marginBottom': '15px'}
                )
            ], style={'width': '48%', 'display': 'inline-block'})
        ]),
        
        html.Div([
            html.Div([
                html.Label("Select Days for First Month:", style={'fontWeight': 'bold', 'marginBottom': '5px', 'display': 'block'}),                dcc.Dropdown(
                    id='month1-days-filter',
                    options=[{'label': day, 'value': day} for day in get_available_days()],
                    value=[],
                    multi=True,
                    placeholder="Select days for first month...",
                    style={'marginBottom': '15px'}
                )
            ], style={'width': '48%', 'display': 'inline-block', 'marginRight': '4%'}),
            
            html.Div([
                html.Label("Select Days for Second Month:", style={'fontWeight': 'bold', 'marginBottom': '5px', 'display': 'block'}),                dcc.Dropdown(
                    id='month2-days-filter',
                    options=[{'label': day, 'value': day} for day in get_available_days()],
                    value=[],
                    multi=True,
                    placeholder="Select days for second month...",
                    style={'marginBottom': '15px'}
                )
            ], style={'width': '48%', 'display': 'inline-block'})
        ]),
        
        html.Div([            html.Div([
                html.Label("Brand Filter:", style={'fontWeight': 'bold', 'marginBottom': '5px', 'display': 'block'}),                dcc.Dropdown(
                    id='brand-filter',
                    options=[{'label': brand, 'value': brand} for brand in get_available_brands()],
                    value=[],
                    multi=True,
                    placeholder="Select brands...",
                    style={'marginBottom': '15px'}
                )
            ], style={'width': '32%', 'display': 'inline-block', 'marginRight': '2%'}),
            
            html.Div([
                html.Label("Category Filter:", style={'fontWeight': 'bold', 'marginBottom': '5px', 'display': 'block'}),                dcc.Dropdown(
                    id='category-filter',
                    options=[{'label': cat, 'value': cat} for cat in get_available_categories()],
                    value=[],
                    multi=True,
                    placeholder="Select categories...",
                    style={'marginBottom': '15px'}
                )
            ], style={'width': '32%', 'display': 'inline-block', 'marginRight': '2%'}),
            
            html.Div([
                html.Label("Day Filter:", style={'fontWeight': 'bold', 'marginBottom': '5px', 'display': 'block'}),                dcc.Dropdown(
                    id='day-filter',
                    options=[{'label': day, 'value': day} for day in get_available_days()],
                    value=[],
                    multi=True,
                    placeholder="Select days...",
                    style={'marginBottom': '15px'}
                )
            ], style={'width': '32%', 'display': 'inline-block'})
        ]),
  
    ], className='filter-section'),
      html.Div([
        html.H3("📊 Category Comparison", 
                style={'marginBottom': '20px', 'color': '#333'}),
        html.Div([
            html.P("📋 Select two different months and adjust filters to see comparison data.", 
                   style={'textAlign': 'center', 'color': '#666', 'fontSize': '16px', 'padding': '20px'})
        ], id='category-comparison-table')
    ], className='comparison-section'),
    
    html.Div([
        html.H3("🛍️ Item Name Comparison", 
                style={'marginBottom': '20px', 'color': '#333'}),
        html.Div([
            html.P("📋 Select two different months and adjust filters to see comparison data.", 
                   style={'textAlign': 'center', 'color': '#666', 'fontSize': '16px', 'padding': '20px'})
        ], id='item-comparison-table')
    ], className='comparison-section')
])

@app.callback(
    [Output('category-comparison-table', 'children'),
     Output('item-comparison-table', 'children')],
    [Input('month1-dropdown', 'value'),
     Input('month2-dropdown', 'value'),
     Input('month1-days-filter', 'value'),
     Input('month2-days-filter', 'value'),
     Input('brand-filter', 'value'),
     Input('category-filter', 'value'),
     Input('day-filter', 'value')]
)
def update_comparison_tables(month1, month2, month1_days, month2_days, brand_filter, category_filter, day_filter):
    if not month1 or not month2 or df.empty:
        return html.Div("Please select both months to compare."), html.Div("Please select both months to compare.")
    
    # Handle None values for filters (convert to empty list)
    month1_days = month1_days or []
    month2_days = month2_days or []
    brand_filter = brand_filter or []
    category_filter = category_filter or []
    day_filter = day_filter or []
    
    # Filter data
    df_month1, df_month2 = filter_data(month1, month2, month1_days, month2_days, brand_filter, category_filter, day_filter)
    
    if df_month1.empty or df_month2.empty:
        return html.Div("No data available for the selected criteria."), html.Div("No data available for the selected criteria.")
    
    # Category Comparison
    category_comparison = create_category_comparison(df_month1, df_month2, month1, month2)
    
    # Item Comparison
    item_comparison = create_item_comparison(df_month1, df_month2, month1, month2)
    
    return category_comparison, item_comparison

def create_category_comparison(df_month1, df_month2, month1, month2):
    """Create category comparison table"""
    try:
        # Ensure Category column is string type and clean data
        df_month1 = df_month1.copy()
        df_month2 = df_month2.copy()
        
        df_month1['Category'] = df_month1['Category'].astype(str)
        df_month2['Category'] = df_month2['Category'].astype(str)
        
        # Filter out invalid categories
        df_month1 = df_month1[df_month1['Category'] != 'nan']
        df_month2 = df_month2[df_month2['Category'] != 'nan']
        
        # Aggregate by category
        agg_month1 = df_month1.groupby('Category')[metrics].sum().reset_index()
        agg_month2 = df_month2.groupby('Category')[metrics].sum().reset_index()
        
        # Merge the data
        comparison = pd.merge(agg_month1, agg_month2, on='Category', suffixes=(f'_{month1}', f'_{month2}'), how='outer').fillna(0)
        
        # Calculate percentage changes
        for metric in metrics:
            col1 = f'{metric}_{month1}'
            col2 = f'{metric}_{month2}'
            comparison[f'{metric}_change'] = comparison.apply(
                lambda row: calculate_percentage_change(row[col1], row[col2]), axis=1
            )        # Create table data
        table_data = []
        for _, row in comparison.iterrows():
            row_data = {'Category': row['Category']}
            
            for metric in metrics:
                col1 = f'{metric}_{month1}'
                col2 = f'{metric}_{month2}'
                change_col = f'{metric}_change'
                
                # Store actual numeric values for sorting, but format for display
                row_data[f'{month1}_{metric}'] = row[col1]
                row_data[f'{month2}_{metric}'] = row[col2]
                row_data[f'Change_{metric}'] = row[change_col] / 100  # Convert to decimal for percentage formatting
            
            table_data.append(row_data)        # Create multi-level columns with proper hierarchy
        columns = [
            {'name': ['', 'Category'], 'id': 'Category', 'type': 'text'}
        ]
        
        for metric in metrics:
            columns.extend([
                {'name': [metric, month1], 'id': f'{month1}_{metric}', 'type': 'numeric', 'format': {'specifier': ',.0f'}},
                {'name': [metric, month2], 'id': f'{month2}_{metric}', 'type': 'numeric', 'format': {'specifier': ',.0f'}},
                {'name': [metric, '% Change'], 'id': f'Change_{metric}', 'type': 'numeric', 'format': {'specifier': '+.1%'}}
            ])          # Style data conditionally for percentage columns
        style_data_conditional = []
        for metric in metrics:
            change_col = f'Change_{metric}'
            # Apply styling based on cell value, not row index
            style_data_conditional.extend([
                {
                    'if': {
                        'filter_query': f'{{{change_col}}} > 0',
                        'column_id': change_col
                    },
                    'color': '#28a745',
                    'fontWeight': 'bold'
                },
                {
                    'if': {
                        'filter_query': f'{{{change_col}}} < 0',
                        'column_id': change_col
                    },
                    'color': '#dc3545',
                    'fontWeight': 'bold'
                },
                {
                    'if': {
                        'filter_query': f'{{{change_col}}} = 0',
                        'column_id': change_col
                    },
                    'color': '#6c757d',
                    'fontWeight': 'bold'
                }
            ])
        return dash_table.DataTable(
            data=table_data,
            columns=columns,
            merge_duplicate_headers=True,
            style_table={'overflowX': 'auto'},
            style_cell={
                'textAlign': 'left',
                'padding': '12px',
                'fontFamily': 'Arial, sans-serif',
                'fontSize': '14px',
                'border': '1px solid #ddd'
            },
            style_header={
                'backgroundColor': '#f8f9fa',
                'fontWeight': 'bold',
                'color': '#333',
                'border': '1px solid #ddd',
                'textAlign': 'center'
            },
            style_data_conditional=style_data_conditional,
            page_size=20,
            sort_action='native',
            sort_mode='multi',
            filter_action='native'
        )
    
    except Exception as e:
        print(f"Error creating category comparison: {e}")
        return html.Div(f"Error creating category comparison table: {str(e)}")

def create_item_comparison(df_month1, df_month2, month1, month2):
    """Create item comparison table"""
    try:
        # Ensure Item name column is string type and clean data
        df_month1 = df_month1.copy()
        df_month2 = df_month2.copy()
        
        df_month1['Item name'] = df_month1['Item name'].astype(str)
        df_month2['Item name'] = df_month2['Item name'].astype(str)
        
        # Filter out invalid item names
        df_month1 = df_month1[df_month1['Item name'] != 'nan']
        df_month2 = df_month2[df_month2['Item name'] != 'nan']
        
        # Aggregate by item name
        agg_month1 = df_month1.groupby('Item name')[metrics].sum().reset_index()
        agg_month2 = df_month2.groupby('Item name')[metrics].sum().reset_index()
        
        # Merge the data
        comparison = pd.merge(agg_month1, agg_month2, on='Item name', suffixes=(f'_{month1}', f'_{month2}'), how='outer').fillna(0)
        
        # Calculate percentage changes
        for metric in metrics:
            col1 = f'{metric}_{month1}'
            col2 = f'{metric}_{month2}'
            comparison[f'{metric}_change'] = comparison.apply(
                lambda row: calculate_percentage_change(row[col1], row[col2]), axis=1
            )
        
        # Sort by total revenue change (descending)
        if 'Item revenue_change' in comparison.columns:
            comparison = comparison.sort_values('Item revenue_change', ascending=False)
        
        # Take top 50 items to avoid overwhelming the table
        comparison = comparison.head(50)        # Create table data
        table_data = []
        for _, row in comparison.iterrows():
            row_data = {'Item name': row['Item name']}
            
            for metric in metrics:
                col1 = f'{metric}_{month1}'
                col2 = f'{metric}_{month2}'
                change_col = f'{metric}_change'
                
                # Store actual numeric values for sorting, but format for display
                row_data[f'{month1}_{metric}'] = row[col1]
                row_data[f'{month2}_{metric}'] = row[col2]
                row_data[f'Change_{metric}'] = row[change_col] / 100  # Convert to decimal for percentage formatting
            
            table_data.append(row_data)          # Create multi-level columns with proper hierarchy
        columns = [
            {'name': ['', 'Item Name'], 'id': 'Item name', 'type': 'text'}
        ]
        
        for metric in metrics:
            columns.extend([
                {'name': [metric, month1], 'id': f'{month1}_{metric}', 'type': 'numeric', 'format': {'specifier': ',.0f'}},
                {'name': [metric, month2], 'id': f'{month2}_{metric}', 'type': 'numeric', 'format': {'specifier': ',.0f'}},
                {'name': [metric, '% Change'], 'id': f'Change_{metric}', 'type': 'numeric', 'format': {'specifier': '+.1%'}}
            ])        
        # Style data conditionally for percentage columns (item comparison)
        style_data_conditional = []
        for metric in metrics:
            change_col = f'Change_{metric}'
            # Apply styling based on cell value, not row index
            style_data_conditional.extend([
                {
                    'if': {
                        'filter_query': f'{{{change_col}}} > 0',
                        'column_id': change_col
                    },
                    'color': '#28a745',
                    'fontWeight': 'bold'
                },
                {
                    'if': {
                        'filter_query': f'{{{change_col}}} < 0',
                        'column_id': change_col
                    },
                    'color': '#dc3545',
                    'fontWeight': 'bold'
                },
                {
                    'if': {
                        'filter_query': f'{{{change_col}}} = 0',
                        'column_id': change_col
                    },
                    'color': '#6c757d',
                    'fontWeight': 'bold'
                }
            ])
        return dash_table.DataTable(
            data=table_data,
            columns=columns,
            merge_duplicate_headers=True,
            style_table={'overflowX': 'auto'},
            style_cell={
                'textAlign': 'left',
                'padding': '12px',
                'fontFamily': 'Arial, sans-serif',
                'fontSize': '14px',
                'border': '1px solid #ddd',
                'maxWidth': '200px',
                'overflow': 'hidden',
                'textOverflow': 'ellipsis'
            },
            style_header={
                'backgroundColor': '#f8f9fa',
                'fontWeight': 'bold',
                'color': '#333',
                'border': '1px solid #ddd',
                'textAlign': 'center'
            },            style_data_conditional=style_data_conditional,
            page_size=20,
            sort_action='native',
            sort_mode='multi',
            filter_action='native',
            tooltip_data=[
                {
                    'Item name': {'value': str(row['Item name']), 'type': 'markdown'}
                } for row in table_data
            ],
            tooltip_duration=None
        )
    
    except Exception as e:
        print(f"Error creating item comparison: {e}")
        return html.Div(f"Error creating item comparison table: {str(e)}")

# Callback to disable child day filter when parent day filters are active
@app.callback(
    [Output('day-filter', 'disabled'),
     Output('day-filter', 'placeholder')],
    [Input('month1-days-filter', 'value'),
     Input('month2-days-filter', 'value')]
)
def disable_child_day_filter(month1_days, month2_days):
    """Disable child day filter when parent day filters are active"""
    # Check if either parent day filter has selections
    month1_has_selection = month1_days and len(month1_days) > 0
    month2_has_selection = month2_days and len(month2_days) > 0
    
    if month1_has_selection or month2_has_selection:
        # Disable child day filter and update placeholder
        return True, "Disabled when month-specific day filters are active"
    else:
        # Enable child day filter with normal placeholder
        return False, "Select days..."

if __name__ == '__main__':
    print("\n" + "="*50)
    print("🚀 Starting Comparison Dashboard...")
    print("="*50)
    
    if df.empty:
        print("❌ No data loaded. Please run mainDashboard.py first to generate cache.")
    else:
        print(f"✅ Data loaded successfully: {df.shape[0]} rows, {df.shape[1]} columns")
        print(f"📅 Available months: {len(get_available_months())}")
        print(f"🏢 Available brands: {len(get_available_brands())}")
        print(f"📦 Available categories: {len(get_available_categories())}")
        
    print("\n🌐 Dashboard will be available at: http://127.0.0.I1:8050")
    print("📊 Features:")
    print("   • Month-to-month comparison")
    print("   • Percentage change calculations with color coding")
    print("   • Filtering by Brand, Category, and Day")
    print("   • Category-wise and Item-wise comparison tables")
    print("   • Sortable and filterable tables")
    print("\n" + "="*50)
    
    app.run(debug=False, host='127.0.0.1', port=8050)
