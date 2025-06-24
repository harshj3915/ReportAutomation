import os
import pandas as pd
import glob
from openpyxl import load_workbook
import time

# Define paths for saved data
cache_folder = os.path.join(os.path.dirname(__file__), 'cache')
cache_file = os.path.join(cache_folder, 'combined_data.pkl')

# Check if cached data exists first, before processing any Excel files
if os.path.exists(cache_file):
    print(f"\nCache file found at {cache_file}")
    print(f"Loading cached data instead of processing Excel files...")
    combined_df = pd.read_pickle(cache_file)
    print("Cached data loaded successfully!")
    
    # Filter out specific categories
    categories_to_exclude = ['JSP', 'Remove', 'FOC-R', 'EMI', '(blank)', 'PRO', 'WRT', 'Others']
    if 'Category' in combined_df.columns:
        print(f"Filtering out categories: {', '.join(categories_to_exclude)}")
        before_count = len(combined_df)
        combined_df = combined_df[~combined_df['Category'].isin(categories_to_exclude)]
        after_count = len(combined_df)
        print(f"Removed {before_count - after_count} rows with excluded categories")
    
    # Process the date column to extract day, month, and year
    if 'Date' in combined_df.columns and not all(col in combined_df.columns for col in ['Day', 'Month', 'Year']):
        print("Processing date column to extract day, month, and year...")
        # Convert to datetime if it's not already
        if not pd.api.types.is_datetime64_dtype(combined_df['Date']):
            combined_df['Date'] = pd.to_datetime(combined_df['Date'], errors='coerce')
        
        # Extract day, month and year
        combined_df['Day'] = combined_df['Date'].dt.day
        combined_df['Month'] = combined_df['Date'].dt.month_name()
        combined_df['Year'] = combined_df['Date'].dt.year
        
        # Save the updated dataframe back to cache
        combined_df.to_pickle(cache_file)
        print("Date columns added and cache updated")
    
    print("\nCombined data shape:", combined_df.shape)
    print("\nCombined data columns:", combined_df.columns.tolist())
    
    # Display the first few rows of the combined data
    print("\nFirst few rows of combined data:")
    print(combined_df.head())
    
else:
    print("No cache found. Processing Excel files...")
    
    # Define the folder path
    folder_path = os.path.join(os.path.dirname(__file__), 'q', 'MainDashboardData')
    print(f"Folder path: {folder_path}")
    
    # Get a list of all Excel files in the folder
    excel_files = glob.glob(os.path.join(folder_path, '*.xlsx'))
    print(f"Found {len(excel_files)} Excel files")
    
    # Function to get the sheet with the longest name from an Excel file
    def get_longest_sheet_name(file_path):
        workbook = load_workbook(file_path, read_only=True)
        sheet_names = workbook.sheetnames
        
        if not sheet_names:
            return None
        
        # Find the sheet with the longest name
        longest_sheet = max(sheet_names, key=len)
        return longest_sheet
    
    # List to store dataframes from each Excel file
    all_dfs = []
    
    # Process each Excel file
    start_time = time.time()
    print("Starting Excel file processing...")
    
    for file in excel_files:
        try:
            # Get the sheet with the longest name
            longest_sheet = get_longest_sheet_name(file)
            
            if longest_sheet:
                print(f"Reading file: {os.path.basename(file)}, Sheet: {longest_sheet}")
                
                # Read the data from the sheet with the longest name
                df = pd.read_excel(file, sheet_name=longest_sheet)
                               
                # Append to our list of dataframes
                all_dfs.append(df)
            else:
                print(f"No sheets found in {os.path.basename(file)}")
                
        except Exception as e:
            print(f"Error processing {os.path.basename(file)}: {str(e)}")
    
    print(f"Excel processing finished in {time.time() - start_time:.2f} seconds")
    
    # Now we have all the data in all_dfs list    # This block only runs if we're processing Excel files (no cache was found)
if 'all_dfs' in locals():
    # Combine all dataframes if we have any
    if all_dfs:
        combined_df = pd.concat(all_dfs, ignore_index=True)
        
        # Filter out specific categories
        categories_to_exclude = ['JSP', 'Remove', 'FOC-R', 'EMI', '(blank)', 'PRO', 'WRT', 'Others']
        if 'Category' in combined_df.columns:
            print(f"Filtering out categories: {', '.join(categories_to_exclude)}")
            before_count = len(combined_df)
            combined_df = combined_df[~combined_df['Category'].isin(categories_to_exclude)]
            after_count = len(combined_df)
            print(f"Removed {before_count - after_count} rows with excluded categories")
        
        # Process the date column to extract day, month, and year
        if 'Date' in combined_df.columns:
            print("Processing date column to extract day, month, and year...")
            # Convert to datetime if it's not already
            if not pd.api.types.is_datetime64_dtype(combined_df['Date']):
                combined_df['Date'] = pd.to_datetime(combined_df['Date'], errors='coerce')
            
            # Extract day, month and year
            combined_df['Day'] = combined_df['Date'].dt.day
            combined_df['Month'] = combined_df['Date'].dt.month_name()
            combined_df['Year'] = combined_df['Date'].dt.year
            print("Date columns added")
        
        print("\nCombined data shape:", combined_df.shape)
        print("\nCombined data columns:", combined_df.columns.tolist())
        
        # Display the first few rows of the combined data
        print("\nFirst few rows of combined data:")
        print(combined_df.head())
        
        # Save the combined dataframe to disk for future use
        try:
            # Create cache folder if it doesn't exist
            if not os.path.exists(cache_folder):
                os.makedirs(cache_folder)
                
            # Save the dataframe
            combined_df.to_pickle(cache_file)
            print(f"\nData saved to {cache_file} for future use")
        except Exception as e:
            print(f"\nError saving data to cache: {str(e)}")
    else:
        print("No data was loaded from the Excel files.")
        combined_df = None